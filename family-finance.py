import pandas as pd
import os
from datetime import datetime
from calendar import month_name
from mapping import CC_REF_NUM, CC_TYPE, CC_TRANS_DATE, CC_CATEGORY, CC_MERCHANT, CC_AMOUNT, B_DEBIT, B_CREDIT, B_BALANCE, B_DATE, B_MERCHANT
from pyautogui import typewrite
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment

# CATEGORIES LIST
categories_exp = ['Books', 'Car Daily', 'Car Maint.', 'Clothing', 'Debt', 'Dining', 'Entertain.',
            'Exercise', 'Gaming', 'Gifts', 'Groceries', 'Hobby', 'Home Decor', 'House Upk.',
            'Insurance', 'Medical', 'Misc.', 'One Time', 'Rent/Mortgage', 'Self Care', 'Trips',
            'Utilities', 'Test']

categories_inc = ['Salary', 'Other Income']

categories_exp = sorted(categories_exp, key=str.lower)
categories_exp_length = len(categories_exp)
categories_inc_length = len(categories_inc)

categories_all = categories_exp + categories_inc
categories_all_length = len(categories_all)

categories_exp.append('Manual')

i = 0
categories_exp_printable = ''
while i < categories_exp_length:
    categories_exp_printable += f"{i+1}. {categories_exp[i]}\n"
    i += 1

i = 0
categories_inc_printable = ''
while i < categories_inc_length:
    categories_inc_printable += f"{i+1}. {categories_inc[i]}\n"
    i += 1

# -- FUNCTIONS --
def choose_new_category(current_expense):
    choice = int(input(categories_exp_printable + f"{categories_exp_length+1}. Manual\nEnter # of desired category for new expense - {current_expense.upper()}: "))
    choice = categories_exp[choice-1]
    return choice

def choose_manual_category(current_expense, amount):
    choice = int(input(categories_exp_printable + f"Enter # of desired category for manual expense - {current_expense.upper()} of {amount}: "))
    choice = categories_exp[choice-1]
    return choice

def choose_new_income_category(current_expense):
    choice = int(input(categories_inc_printable + f"Enter # of desired category for new income - {current_expense.upper()}: "))
    choice = categories_inc[choice-1]
    return choice

def category_to_int(choice):
    index = categories_all.index(choice)
    return index + 1

def populate_ec(current_expense, category):
    sf_sheet_ec.cell(column=1, row=j+2).value = current_expense
    sf_sheet_ec.cell(column=2, row=j+2).value = category

def populate_se(current_expense, category, amount):
    if category in categories_inc:
        sf_sheet_se.cell(column=5, row=i+2).value = current_expense
        sf_sheet_se.cell(column=6, row=i+2).value = category
        sf_sheet_se.cell(column=7, row=i+2).value = amount
    else:
        sf_sheet_se.cell(column=1, row=i+2).value = current_expense
        sf_sheet_se.cell(column=2, row=i+2).value = category
        sf_sheet_se.cell(column=3, row=i+2).value = amount

def populate_ec_se(current_expense, category, amount):
    populate_ec(current_expense, category)
    populate_se(current_expense, category, amount)

def trim_merchant_name(current_expense):
    print("Would you like to adjust the merchant name?")
    typewrite(current_expense.upper())
    current_expense = input().strip()
    return current_expense

# Change .csv file to .xlsx to allow openpyxl to work with data
def convert_csv(file_name):
    read_path = file_name + '.csv'
    write_path = file_name + '.xlsx'
    cwd = os.getcwd()
    read_file = pd.read_csv (os.path.join(cwd, read_path))
    read_file.to_excel (os.path.join(cwd, write_path), index=None, header=True)

def calculate_average_cd(col):
    i = 0
    total = 0
    blank_cells = 0
    while i < recorded_length:
        if sf_sheet_cd.cell(column=col, row=i+3).value == None:
            blank_cells += 1
        else:
            total += float(sf_sheet_cd.cell(column=col, row=i+3).value)
        i += 1
    if recorded_length - blank_cells == 0:
        return 0
    else:
        return total / (recorded_length - blank_cells)

def format_ieal(sc, sr, worksheet):
    sc -= 1
    sr -= 1
    ws = sf_workbook[worksheet]

    # Merge cells for Title, Income, Expense, Assets, Liabilities, Liquid Inv. Info, Actual, Actual
    ws.merge_cells(start_column=sc+1, start_row=sr+1, end_column=sc+13, end_row=sr+1)
    ws.merge_cells(start_column=sc+1, start_row=sr+2, end_column=sc+2, end_row=sr+2)
    ws.merge_cells(start_column=sc+3, start_row=sr+2, end_column=sc+4, end_row=sr+2)
    ws.merge_cells(start_column=sc+7, start_row=sr+2, end_column=sc+8, end_row=sr+2)
    ws.merge_cells(start_column=sc+9, start_row=sr+2, end_column=sc+10, end_row=sr+2)
    ws.merge_cells(start_column=sc+12, start_row=sr+2, end_column=sc+13, end_row=sr+2)
    ws.merge_cells(start_column=sc+1, start_row=sr+3, end_column=sc+2, end_row=sr+3)
    ws.merge_cells(start_column=sc+3, start_row=sr+3, end_column=sc+4, end_row=sr+3)

    # Populate text for Title, Income, Expense, Assets, Liabilities, Liquid Inv. Info, Actual, Actual
    ws.cell(column=sc+1, row=sr+1).value = f"{month_name[desired_month]} {desired_year}"
    ws.cell(column=sc+1, row=sr+2).value = "Income"
    ws.cell(column=sc+3, row=sr+2).value = "Expense"
    ws.cell(column=sc+7, row=sr+2).value = "Assets"
    ws.cell(column=sc+9, row=sr+2).value = "Liabilities"
    ws.cell(column=sc+12, row=sr+2).value = "Liquid Inv. Info"
    ws.cell(column=sc+1, row=sr+3).value = "Actual"
    ws.cell(column=sc+3, row=sr+3).value = "Actual"

    # Populate text for Cash, Joint, Liquid Inv., Retirement, Joe, Jane, Investments, Transferred in, Current Value, Return
    ws.cell(column=sc+7, row=sr+3).value = "Cash"
    ws.cell(column=sc+7, row=sr+4).value = "Joint"
    ws.cell(column=sc+7, row=sr+5).value = "Liquid Inv."
    ws.cell(column=sc+7, row=sr+7).value = "Retirement"
    ws.cell(column=sc+7, row=sr+8).value = "Joe"
    ws.cell(column=sc+7, row=sr+9).value = "Jane"
    ws.cell(column=sc+7, row=sr+10).value = "Investments"
    ws.cell(column=sc+12, row=sr+4).value = "Transferred in"
    ws.cell(column=sc+12, row=sr+5).value = "Current Value"
    ws.cell(column=sc+12, row=sr+6).value = "Return"

    # Populate text and amounts for Income categories
    i = 0
    while i < categories_inc_length:
        ws.cell(column=sc+1, row=sr+4+i).value = income_list_sorted[i]['category']
        ws.cell(column=sc+2, row=sr+4+i).value = income_list_sorted[i]['amount']
        i += 1

    ws.cell(column=sc+2, row=sr+categories_exp_length+4).value = income_total

    # Populate text and amounts for Expense categories
    i = 0
    while i < categories_exp_length:
        ws.cell(column=sc+3, row=sr+4+i).value = expense_list_sorted[i]['category']
        ws.cell(column=sc+4, row=sr+4+i).value = expense_list_sorted[i]['amount']
        i += 1

    ws.cell(column=sc+4, row=sr+categories_exp_length+4).value = expense_total
    ws.cell(column=sc+5, row=sr+categories_exp_length+4).value = income_total - expense_total

    # Populate amounts for Assets
    ws.cell(column=sc+8, row=sr+4).value = checking_bal
    ws.cell(column=sc+8, row=sr+5).value = investment_bal
    ws.cell(column=sc+8, row=sr+6).value = cash_total
    ws.cell(column=sc+8, row=sr+8).value = joe_retirement_bal
    ws.cell(column=sc+8, row=sr+9).value = jane_retirement_bal
    ws.cell(column=sc+8, row=sr+10).value = other_retirement_bal
    ws.cell(column=sc+8, row=sr+11).value = assets_total

    # Populate amount for Liabilities
    ws.cell(column=sc+10, row=sr+11).value = liabilities_total

    # Populate amount for Net Worth
    ws.cell(column=sc+11, row=sr+11).value = assets_total - liabilities_total

    # Populate amounts for Liquid Inv. Info
    investment_return = investment_bal - investment_transfer_total
    ws.cell(column=sc+13, row=sr+4).value = investment_transfer_total
    ws.cell(column=sc+13, row=sr+5).value = investment_bal
    ws.cell(column=sc+13, row=sr+6).value = investment_return
    if investment_transfer_total == 0:
        ws.cell(column=sc+13, row=sr+7).value = "N/A"
    else:
        ws.cell(column=sc+13, row=sr+7).value = investment_return / investment_transfer_total

    # Generic Styles
    medium = Side(border_style="medium", color="000000")
    thin = Side(border_style="thin", color="000000")

    # Style label for Title
    ws.cell(column=sc+1, row=sr+1).font = Font(bold=True)
    ws.cell(column=sc+1, row=sr+1).alignment = Alignment(horizontal="center")
    if sr > 1:
        i = 0
        while i < 13:
            ws.cell(column=sc+1+i, row=sr+1).border = Border(top=medium)
            i += 1

    # Style labels for Income, Expense, Assets, Liabilities, Liquid Inv. Info, Actual, Actual
    ws.cell(column=sc+1, row=sr+2).alignment = Alignment(horizontal="center")
    ws.cell(column=sc+3, row=sr+2).alignment = Alignment(horizontal="center")
    ws.cell(column=sc+7, row=sr+2).alignment = Alignment(horizontal="center")
    ws.cell(column=sc+9, row=sr+2).alignment = Alignment(horizontal="center")
    ws.cell(column=sc+12, row=sr+2).alignment = Alignment(horizontal="center")
    ws.cell(column=sc+1, row=sr+3).alignment = Alignment(horizontal="center")
    ws.cell(column=sc+3, row=sr+3).alignment = Alignment(horizontal="center")

    # Style labels for Joint, Liquid Inv., Joe, Jane, Return
    ws.cell(column=sc+7, row=sr+4).alignment = Alignment(indent=1)
    ws.cell(column=sc+7, row=sr+5).alignment = Alignment(indent=1)
    ws.cell(column=sc+7, row=sr+8).alignment = Alignment(indent=1)
    ws.cell(column=sc+7, row=sr+9).alignment = Alignment(indent=1)
    ws.cell(column=sc+12, row=sr+6).alignment = Alignment(indent=1)

    # Style number format for Incomes
    ws.cell(column=sc+2, row=sr+4).style = 'Currency'

    i = 0
    while i < categories_exp_length - 1:
        ws.cell(column=sc+2, row=sr+5+i).style = 'Comma'
        i += 1

    ws.cell(column=sc+2, row=sr+categories_exp_length+4).style = 'Currency'

    # Style number format for Expenses
    ws.cell(column=sc+4, row=sr+4).style = 'Currency'

    i = 0
    while i < categories_exp_length - 1:
        ws.cell(column=sc+4, row=sr+5+i).style = 'Comma'
        i += 1

    ws.cell(column=sc+4, row=sr+categories_exp_length+4).style = 'Currency'
    ws.cell(column=sc+5, row=sr+categories_exp_length+4).style = 'Currency'
    ws.cell(column=sc+5, row=sr+categories_exp_length+4).font = Font(bold=True)

    # Stlye number format for Assets and Liabilities
    i = 0
    while i < 2:
        ws.cell(column=sc+8+i*2, row=sr+4).style = 'Currency'
        ws.cell(column=sc+8+i*2, row=sr+5).style = 'Comma'
        ws.cell(column=sc+8+i*2, row=sr+6).style = 'Currency'

        ws.cell(column=sc+8+i*2, row=sr+8).style = 'Comma'
        ws.cell(column=sc+8+i*2, row=sr+9).style = 'Comma'
        ws.cell(column=sc+8+i*2, row=sr+10).style = 'Comma'
        ws.cell(column=sc+8+i*2, row=sr+11).style = 'Currency'
        i += 1

    ws.cell(column=sc+11, row=sr+11).style = 'Currency'
    ws.cell(column=sc+11, row=sr+11).font = Font(bold=True)

    # Style number format for Liquid Inv.
    ws.cell(column=sc+13, row=sr+4).style = 'Currency'
    ws.cell(column=sc+13, row=sr+5).style = 'Comma'
    ws.cell(column=sc+13, row=sr+6).style = 'Currency'
    ws.cell(column=sc+13, row=sr+7).style = 'Percent'
    ws.cell(column=sc+13, row=sr+7).number_format = '0.00%'

    # Style border for Income, Expense, Assets, Liabilities, Liquid Inv.
    ws.cell(column=sc+2, row=sr+categories_exp_length+4).border = Border(top=thin)
    ws.cell(column=sc+4, row=sr+categories_exp_length+4).border = Border(top=thin)
    ws.cell(column=sc+8, row=sr+6).border = Border(top=thin)
    ws.cell(column=sc+8, row=sr+11).border = Border(top=thin)
    ws.cell(column=sc+10, row=sr+11).border = Border(top=thin)
    ws.cell(column=sc+13, row=sr+6).border = Border(top=thin)

# -- END FUNCTIONS --
convert_csv(r'ccd')
convert_csv(r'bd')

# Load workbooks
cc_data = load_workbook(filename="ccd.xlsx")
cc_sheet = cc_data.active
b_data = load_workbook(filename="bd.xlsx")
b_sheet = b_data.active
sf_workbook = load_workbook(filename="FamilyFinance.xlsx")
sf_sheet_rd = sf_workbook["RD"] # Raw Data
sf_sheet_ec = sf_workbook["EC"] # Expense - Category
sf_sheet_se = sf_workbook["SE"] # Sorted Expenses
sf_sheet_ce = sf_workbook["CE"] # Categorized Expenses
sf_sheet_cd = sf_workbook["CD"] # Chart Data
sf_sheet_overview = sf_workbook["Overview"] # Overview
sf_sheet_hd = sf_workbook["Historical Data"] # Historical Data
sf_sheet_charts = sf_workbook["Charts"] # Charts


# Clear prior data from RD, SE, CE, Overview
sf_sheet_rd.delete_rows(idx=2, amount=300)
sf_sheet_se.delete_rows(idx=2, amount=300)
sf_sheet_ce.delete_rows(idx=3, amount=300)
sf_sheet_overview.delete_rows(idx=2, amount=300)

# Populate categories for CE and CD
i = 0
while i < categories_all_length:
    sf_sheet_ce.cell(column=((i+1)*2)-1, row=1).value = categories_all[i]
    sf_sheet_cd.cell(column=15+i, row=1).value = categories_all[i]
    i += 1

# Declare lists
transactions = []

# Format needed info from ccd.xlsx
for row in cc_sheet.iter_rows(min_row=2, values_only=True):
    raw_date = row[CC_TRANS_DATE]
    parsed_date = datetime.strptime(raw_date, "%m/%d/%Y")
    raw_amount = row[CC_AMOUNT].strip()
    raw_amount = raw_amount.replace(',','')
    raw_amount = raw_amount.replace('(','')
    raw_amount = raw_amount.replace(')','')
    raw_amount = raw_amount.replace('$','')

    trans = {
        "ref_num": row[CC_REF_NUM].strip(),
        "type": row[CC_TYPE].strip(),
        "trans_date": parsed_date,
        "trans_year": parsed_date.year,
        "trans_month": parsed_date.month,
        "trans_day": parsed_date.day,
        "category": row[CC_CATEGORY].strip(),
        "merchant": row[CC_MERCHANT].strip(),
        "balance": None,
        "amount": float(raw_amount)
    }
    transactions.append(trans)

# Determine what should be taken from cc_transactions & b_transactions
desired_year = int(input("Enter desired year in 20XX format: "))
desired_month = int(input("Enter desired month in XX format: "))

# Filter through cc_transactions, removing purchases outside of desired month and all credit card payments
transactions = list(filter(lambda trans: trans["trans_month"] == desired_month, transactions))
transactions = list(filter(lambda trans: trans["trans_year"] == desired_year, transactions))
transactions = list(filter(lambda trans: trans["type"] == 'Purchase', transactions))

# Format needed info from bd.xlsx
for row in b_sheet.iter_rows(min_row=2, values_only=True):
    raw_date = row[B_DATE]
    parsed_date = datetime.strptime(raw_date, "%m/%d/%Y")
    if row[B_DEBIT] == None:
        trans = {
            "ref_num": None,
            "type": 'Income',
            "trans_date": parsed_date,
            "trans_year": parsed_date.year,
            "trans_month": parsed_date.month,
            "trans_day": parsed_date.day,
            "category": None,
            "merchant": row[B_MERCHANT].strip(),
            "balance": row[B_BALANCE],
            "amount": row[B_CREDIT]
        }
        if 'JOHN DOE CORP' in trans['merchant'] and trans['trans_day'] > 15:
            trans['trans_month'] - 1
            trans['trans_day'] = 3
        transactions.append(trans)
    else:
        if 'CREDIT CARD' in row[B_MERCHANT].upper():
            None
        elif 'TRANSFER TO X' in row[B_MERCHANT].upper():
            None
        else:
            trans = {
                "ref_num": None,
                "type": 'Purchase',
                "trans_date": parsed_date,
                "trans_year": parsed_date.year,
                "trans_month": parsed_date.month,
                "trans_day": parsed_date.day,
                "category": None,
                "merchant": row[B_MERCHANT].strip(),
                "balance": None,
                "amount": row[B_DEBIT]
            }
            transactions.append(trans)

# Filter through all transactions
transactions = list(filter(lambda trans: trans["trans_month"] == desired_month, transactions))
transactions = list(filter(lambda trans: trans["trans_year"] == desired_year, transactions))

# Populate raw data RD in FamilyFinance.xlsx
trans_length = len(transactions)
i = 0
while i < trans_length:
    sf_sheet_rd.cell(column=1, row=i+2).value = transactions[i]["ref_num"]
    sf_sheet_rd.cell(column=2, row=i+2).value = transactions[i]["type"]
    sf_sheet_rd.cell(column=3, row=i+2).value = transactions[i]["trans_date"]
    sf_sheet_rd.cell(column=4, row=i+2).value = transactions[i]["trans_year"]
    sf_sheet_rd.cell(column=5, row=i+2).value = transactions[i]["trans_month"]
    sf_sheet_rd.cell(column=6, row=i+2).value = transactions[i]["trans_day"]
    sf_sheet_rd.cell(column=7, row=i+2).value = transactions[i]["category"]
    sf_sheet_rd.cell(column=8, row=i+2).value = transactions[i]["merchant"]
    sf_sheet_rd.cell(column=9, row=i+2).value = transactions[i]["balance"]
    sf_sheet_rd.cell(column=10, row=i+2).value = transactions[i]["amount"]
    i += 1

# Categorize transactions
i = 0
while i < trans_length:
    current_expense = sf_sheet_rd.cell(column=8, row=i+2).value
    current_amount = sf_sheet_rd.cell(column=10, row=i+2).value
    current_type = sf_sheet_rd.cell(column=2, row=i+2).value

    # Find how many listed merchants there are with matching categories
    merchant_length = -1
    merchant_len_finder = 0

    j = 0
    while merchant_len_finder != None:
        merchant_len_finder = sf_sheet_ec.cell(column=1, row=j+2).value
        merchant_length += 1
        j += 1
    
    # Check if a categorized merchant matches the expense being tested
    j = 0
    current_expense_category = ''
    while j < merchant_length:
        if sf_sheet_ec.cell(column=1, row=j+2).value.upper() in current_expense.upper():
            current_expense_category = sf_sheet_ec.cell(column=2, row=j+2).value
            current_expense = sf_sheet_ec.cell(column=1, row=j+2).value
            break
        else:
            j += 1

    # If the current expense isn't found, add it to the EC list and SE list
    if j == merchant_length:
        if current_type == 'Purchase':
            new_category = choose_new_category(current_expense)

            current_expense = trim_merchant_name(current_expense)

            if new_category == 'Manual':
                chosen_category = choose_manual_category(current_expense, current_amount)
                populate_ec(current_expense, new_category)
                populate_se(current_expense, chosen_category, current_amount)
            else:
                populate_ec_se(current_expense, new_category, current_amount)
        else:
            new_category = choose_new_income_category(current_expense)

            current_expense = trim_merchant_name(current_expense)

            populate_ec_se(current_expense, new_category, current_amount)
    
    # If the found category is 'Manual', choose category and add to SE list - Else populate with found category and add to SE list
    elif current_expense_category == 'Manual':
        chosen_category = choose_manual_category(current_expense, current_amount)

        populate_se(current_expense, chosen_category, current_amount)

    # Add the current expense to the SE with its matching category
    else:
        if current_type == 'Purchase':
            populate_se(current_expense, current_expense_category, current_amount)
        else:
            populate_se(current_expense, current_expense_category, current_amount)

    i += 1

# Sort by categories
i = 0
while i < trans_length:
    if sf_sheet_se.cell(column=2, row=i+2).value == None:
        current_expense = sf_sheet_se.cell(column=5, row=i+2).value
        current_expense_category = sf_sheet_se.cell(column=6, row=i+2).value
        current_amount = sf_sheet_se.cell(column=7, row=i+2).value
    else:
        current_expense = sf_sheet_se.cell(column=1, row=i+2).value
        current_expense_category = sf_sheet_se.cell(column=2, row=i+2).value
        current_amount = sf_sheet_se.cell(column=3, row=i+2).value

    ce_column_category = int(category_to_int(current_expense_category))*2 - 1
    ce_column_amount =int(category_to_int(current_expense_category))*2

    sf_sheet_ce.cell(column=ce_column_category, row=i+3).value = current_expense
    sf_sheet_ce.cell(column=ce_column_amount, row=i+3).value = current_amount

    i += 1

i = 0
column_expense_total = 0
while i < len(categories_all):
    j = 0
    while j < trans_length:
        if sf_sheet_ce.cell(column=(i+1)*2, row=j+3).value != None:
            column_expense_total += sf_sheet_ce.cell(column=(i+1)*2, row=j+3).value
        j += 1
    sf_sheet_ce.cell(column=(i+1)*2, row=2).value = column_expense_total
    column_expense_total = 0
    i += 1

# Determine existing rows
recorded_length = 0
i = 0
j = 0
while i < 1:
    if sf_sheet_cd.cell(column=1, row=j+3).value != None:
        recorded_length += 1
        j += 1
    else:
        i = 1

# Find which row will be populated
working_row = 0
i = 0
while i < recorded_length:
    if sf_sheet_cd.cell(column=2, row=i+3).value == desired_year and sf_sheet_cd.cell(column=1, row=i+3).value == desired_month:
        working_row = i+3
        break
    else:
        i += 1

if i == recorded_length:
    working_row = i+3

# Populate working_row with data
i = 0
expense_total = 0
income_total = 0
current_expense_value = 0
while i < len(categories_all):
    current_expense_value = float(sf_sheet_ce.cell(column=(i+1)*2, row=2).value)
    sf_sheet_cd.cell(column=i+15, row=working_row).value = current_expense_value
    if i < categories_exp_length:
        expense_total += current_expense_value
    else:
        income_total += current_expense_value
    i += 1

# Find most recent transaction and if it has an associated balance
i = 0
most_recent_date = datetime(1, 1, 1)
most_recent_date_row = 1
while i < trans_length:
    if sf_sheet_rd.cell(column=3, row=i+2).value > most_recent_date and sf_sheet_rd.cell(column=9, row=i+2).value != None:
        most_recent_date = sf_sheet_rd.cell(column=3, row=i+2).value
        most_recent_date_row = i+2
    i += 1

# Collect misc. information
if sf_sheet_rd.cell(column=9, row=most_recent_date_row).value != None:
    checking_bal = float(sf_sheet_rd.cell(column=9, row=most_recent_date_row).value)
    print(f"Checking Account balance automatically pulled as {checking_bal}")
else:
    checking_bal = float(input("Please input end of month Checking Account balance: "))

investment_bal = float(input("Please input end of month Investment Account balance: "))
investment_transfer_current = float(input("Please input how much was transferred to Investment Account during the month: "))
cash_total = checking_bal + investment_bal
joe_retirement_bal = float(input("Please input Joe's end of month Retirement Account balance: "))
jane_retirement_bal = float(input("Please input Janes's end of month Retirement Account balance: "))
other_retirement_bal = float(input("Please input other end of month Retirement Account balances total: "))
retirement_total = joe_retirement_bal + jane_retirement_bal
assets_total = cash_total + retirement_total
liabilities_total = float(input("Please input Total Liabilities: "))

# Populate CD with info
sf_sheet_cd.cell(column=1, row=working_row).value = desired_month
sf_sheet_cd.cell(column=2, row=working_row).value = desired_year
sf_sheet_cd.cell(column=3, row=working_row).value = f"=DATE(B{working_row},A{working_row},1)"
sf_sheet_cd.cell(column=3, row=working_row).number_format = 'mmm-yy'

sf_sheet_cd.cell(column=4, row=working_row).value = income_total
sf_sheet_cd.cell(column=5, row=working_row).value = expense_total
sf_sheet_cd.cell(column=6, row=working_row).value = income_total - expense_total
sf_sheet_cd.cell(column=7, row=working_row).value = cash_total
sf_sheet_cd.cell(column=8, row=working_row).value = retirement_total
sf_sheet_cd.cell(column=9, row=working_row).value = other_retirement_bal
sf_sheet_cd.cell(column=10, row=working_row).value = assets_total
sf_sheet_cd.cell(column=11, row=working_row).value = liabilities_total
sf_sheet_cd.cell(column=12, row=working_row).value = assets_total - liabilities_total
if sf_sheet_cd.cell(column=13, row=working_row-1).value == None:
    investment_transfer_prior = 0
else:
    investment_transfer_prior = sf_sheet_cd.cell(column=13, row=working_row-1).value
investment_transfer_total = investment_transfer_prior + investment_transfer_current
sf_sheet_cd.cell(column=13, row=working_row).value = investment_transfer_total

i = 0
while i < 100:
    sf_sheet_cd.cell(column=4+i, row=working_row).style = 'Comma'
    i += 1

# Populate Averages in CD
i = 0
while i < 3:
    sf_sheet_cd.cell(column=4+i, row=2).value = calculate_average_cd(4+i)
    i += 1

i = 0
while i < len(categories_all):
    sf_sheet_cd.cell(column=15+i, row=2).value = calculate_average_cd(15+i)
    i += 1

# Create lists from CD info and create a separate sorted lists
i = 0
expense_list = []
while i < categories_exp_length:
    expense_cat = sf_sheet_cd.cell(column=i+15, row=1).value
    expense_amount = sf_sheet_cd.cell(column=i+15, row=working_row).value
    expense = {
        "category": expense_cat,
        "amount": expense_amount
        }
    expense_list.append(expense)
    i += 1

i = 0
income_list = []
while i < categories_inc_length:
    income_cat = sf_sheet_cd.cell(column=i+15+categories_exp_length, row=1).value
    income_amount = sf_sheet_cd.cell(column=i+15+categories_exp_length, row=working_row).value
    income = {
        "category": income_cat,
        "amount": income_amount
    }
    income_list.append(income)
    i += 1

expense_list_sorted = sorted(expense_list, key = lambda i: i["amount"], reverse=True)
income_list_sorted = sorted(income_list, key = lambda i: i["amount"], reverse=True)

# Populate Overview info for active month
format_ieal(1, 1, "Overview")

# Populate Historical Data info for active month
hd_row = ((desired_year - 2021) * (categories_exp_length+4) * 12) + (desired_month * (categories_exp_length+4)) + 1
format_ieal(1, hd_row, "Historical Data")

# Save FamilyFinance.xlsx with updated information
sf_workbook.save(filename="FamilyFinance.xlsx")